#!/usr/bin/env python3
"""
Turn the three regional shipment trackers into the cockpit's dataset.

    python3 tools/import_trackers.py PK.xlsx BD.xlsx SL.xlsx -o src/dataset.json

Each workbook is a SAP HANA export with a per-region sheet (MSBD / PK-REGION /
SL_REGION). The sheet's header sits on row 4; data starts on row 5.

The trackers are maintained by hand across three sites, so the same status is
spelled several ways and date columns carry text. Everything this script
normalises is counted and reported, so the cleanup is visible rather than
silent. NOTE: the output contains customer names, prices and order values —
keep it out of version control.
"""

import argparse
import datetime
import json
import re
import sys
from collections import Counter, defaultdict

import openpyxl

HEADER_ROW = 4

# Which sheet holds the tracker, per source region.
SHEETS = {
    'PK': ('PK-REGION',),
    'BD': ('MSBD', 'MS-BD'),
    'SL': ('SL_REGION', 'SL-REGION'),
}

REGION_NAMES = {'PK': 'Pakistan', 'BD': 'Bangladesh', 'SL': 'Sri Lanka'}

# Header labels vary by a space or a newline between files, so match loosely.
FIELDS = {
    'id': 'ID',
    'punched': 'Line Punched',
    'planMonth': 'Plan Month',
    'plant': 'Plant',
    'customer': 'Customer',
    'destination': 'Unloading Point',
    'salesRegion': 'Region',
    'aam': 'AAM',
    'kam': 'KAM',
    'userStatus': 'User Status',
    'salesOrder': 'SD',
    'lineItem': 'SD Item',
    'confQty': 'Conf Qty',
    'remQty': 'Rem Qty',
    'asp': 'ASP',
    'confValue': 'Conf Value',
    'balToShip': 'Bal to Ship',
    'confirmed': 'Conf Del Date',
    'shipMode': 'Ship Mode',
    'shipPlanStatus': 'Status_ship_Plan',
    'shipPlanDate': 'Date_Ship_Plan',
    'agentStatus': 'Status_Agent_Details',
    'agentDate': 'Date_Agent',
    'bookingStatus': 'Booking Status',
    'bookingDate': 'Booking Date',
    'croStatus': 'CRO/LP Status',
    'croDate': 'CRO/LP Date',
    'readiness': 'Good Readiness Status',
    'remarks': 'Remarks/Challenge',
}

# Text that means "no value" in a column that should hold a date.
NULLISH = {'', '-', '--', 'n/a', 'na', '#n/a', '#ref!', '#value!', 'pending', 'tbd', '.'}

# Booking is "Placed" in Pakistan and "Received" elsewhere for the same state.
BOOKING_SYNONYMS = {'placed': 'Received', 'received': 'Received', 'pending': 'Pending'}

report = defaultdict(Counter)


def norm_key(label):
    """Collapse whitespace/newlines so 'CRO/LP  Status' == 'CRO/LP Status'."""
    return re.sub(r'\s+', ' ', str(label or '')).strip().lower()


def clean_status(value, region, column):
    """Trim and title-case a status; count every value that had to be changed."""
    raw = '' if value is None else str(value).strip()
    if norm_key(raw) in NULLISH:
        return ''
    fixed = raw[:1].upper() + raw[1:] if raw else raw
    fixed = ' '.join(w.capitalize() if w.isupper() is False and w.islower() else w
                     for w in fixed.split())
    if fixed != raw:
        report[f'case/{column}'][f'{region}: {raw!r} -> {fixed!r}'] += 1
    return fixed


def as_date(value, region, column):
    """Return a date, or None — counting the text found in date columns."""
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = '' if value is None else str(value).strip()
    if norm_key(text) in NULLISH:
        if text:
            report[f'text-in-date/{column}'][f'{region}: {text!r}'] += 1
        return None
    report[f'text-in-date/{column}'][f'{region}: {text!r}'] += 1
    return None


def num(value):
    return round(float(value), 2) if isinstance(value, (int, float)) else 0.0


def pick_sheet(book, region):
    for name in SHEETS[region]:
        if name in book.sheetnames:
            return book[name]
    raise SystemExit(f'{region}: none of {SHEETS[region]} found in {book.sheetnames}')


def detect_region(book, path):
    for region, names in SHEETS.items():
        if any(n in book.sheetnames for n in names):
            return region
    raise SystemExit(f'cannot tell which region {path} is — sheets: {book.sheetnames}')


def read_tracker(path):
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    region = detect_region(book, path)
    sheet = pick_sheet(book, region)

    rows = sheet.iter_rows(min_row=HEADER_ROW, values_only=True)
    header = [norm_key(h) for h in next(rows)]
    index = {}
    for key, label in FIELDS.items():
        want = norm_key(label)
        index[key] = header.index(want) if want in header else None
        if index[key] is None:
            report['missing-column'][f'{region}: {label}'] += 1

    def cell(row, key):
        i = index[key]
        return row[i] if i is not None and i < len(row) else None

    lines = []
    for row in rows:
        if not row or row[0] in (None, ''):
            continue
        rec = {'region': region, 'regionName': REGION_NAMES[region]}
        for key in FIELDS:
            rec[key] = cell(row, key)
        lines.append(rec)

    book.close()
    return region, lines


def build(records, today):
    """Normalise, derive the cockpit's status model, and drop empty columns."""
    out = []
    for rec in records:
        region = rec['region']

        agent = clean_status(rec['agentStatus'], region, 'Status_Agent_Details')
        cro = clean_status(rec['croStatus'], region, 'CRO/LP Status')
        booking_raw = clean_status(rec['bookingStatus'], region, 'Booking Status')
        booking = BOOKING_SYNONYMS.get(booking_raw.lower(), booking_raw)
        if booking_raw and booking != booking_raw:
            report['vocabulary/Booking Status'][f'{region}: {booking_raw!r} -> {booking!r}'] += 1

        punched = as_date(rec['punched'], region, 'Line Punched')
        confirmed = as_date(rec['confirmed'], region, 'Conf Del Date')
        agent_date = as_date(rec['agentDate'], region, 'Date_Agent')
        cro_date = as_date(rec['croDate'], region, 'CRO/LP Date')
        booking_date = as_date(rec['bookingDate'], region, 'Booking Date')

        # A line ages from its last completed milestone; if nothing has
        # completed, from the day it was punched.
        since = max([d for d in (booking_date, agent_date, punched) if d], default=None)
        aging = (today - since).days if since else 0

        agent_ok = agent == 'Received'
        cro_ok = cro == 'Received'
        if agent_ok and cro_ok:
            overall = 'READY'
        elif agent_ok:
            overall = 'CRO PENDING'
        elif cro_ok:
            overall = 'AGENT DETAILS PENDING'
        else:
            overall = 'ACTION REQUIRED'

        out.append({
            'id': str(rec['id']),
            'region': region,
            'regionName': REGION_NAMES[region],
            'plant': str(rec['plant'] or ''),
            'customer': str(rec['customer'] or '').strip(),
            'salesRegion': str(rec['salesRegion'] or '').strip(),
            'destination': str(rec['destination'] or '').strip(),
            'owner': str(rec['aam'] or '').strip() or 'Unassigned',
            'kam': str(rec['kam'] or '').strip(),
            'salesOrder': str(rec['salesOrder'] or ''),
            'lineItem': str(rec['lineItem'] or '').lstrip('0') or '0',
            'planMonth': str(rec['planMonth'] or '').strip(),
            'userStatus': str(rec['userStatus'] or '').strip(),
            'shipMode': str(rec['shipMode'] or '').strip(),
            'confQty': num(rec['confQty']),
            'remQty': num(rec['remQty']),
            'asp': num(rec['asp']),
            'confValue': num(rec['confValue']),
            'balToShip': num(rec['balToShip']),
            'agentStatus': agent or 'Pending',
            'agentDate': agent_date.isoformat() if agent_date else '',
            'bookingStatus': booking or 'Pending',
            'bookingDate': booking_date.isoformat() if booking_date else '',
            'croStatus': f'CRO {cro}' if cro else 'CRO Pending',
            'croDate': cro_date.isoformat() if cro_date else '',
            'confirmed': confirmed.isoformat() if confirmed else '',
            'daysToDelivery': (confirmed - today).days if confirmed else None,
            'punched': punched.isoformat() if punched else '',
            'aging': aging,
            'overall': overall,
            'readiness': str(rec['readiness'] or '').strip(),
            'remarks': str(rec['remarks'] or '').strip()[:180],
        })
    return out


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('workbooks', nargs='+')
    ap.add_argument('-o', '--out', default='src/dataset.json')
    ap.add_argument('--today', help='extract date (YYYY-MM-DD); defaults to the latest line punched')
    args = ap.parse_args()

    records = []
    sources = []
    for path in args.workbooks:
        region, rows = read_tracker(path)
        sources.append({'region': region, 'file': path.split('/')[-1], 'rows': len(rows)})
        print(f'  {region}: {len(rows):>5} lines from {path.split("/")[-1]}', file=sys.stderr)
        records.extend(rows)

    if args.today:
        today = datetime.date.fromisoformat(args.today)
    else:
        punched = [as_date(r['punched'], r['region'], 'Line Punched') for r in records]
        today = max(d for d in punched if d)

    lines = build(records, today)

    payload = {
        'generatedFrom': 'regional shipment trackers',
        'asOf': today.isoformat(),
        'sources': sources,
        'lines': lines,
    }
    with open(args.out, 'w', encoding='utf-8') as fh:
        json.dump(payload, fh, separators=(',', ':'))

    print(f'\n{len(lines)} lines -> {args.out}  (as of {today})', file=sys.stderr)
    if report:
        print('\nNormalised on the way in:', file=sys.stderr)
        for kind, counter in sorted(report.items()):
            total = sum(counter.values())
            print(f'  {kind}: {total} rows', file=sys.stderr)
            for what, n in counter.most_common(6):
                print(f'      {n:>5}  {what}', file=sys.stderr)


if __name__ == '__main__':
    main()
